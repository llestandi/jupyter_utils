# Import necessary libraries
import pandas as pd
from icalendar import Calendar, Event
from datetime import datetime, time
from datetime import datetime, timedelta
import calendar
import pytz
import openpyxl
import re

def read_ics(file_path: str):
    """
    Reads and parses an ICS (iCalendar) file, extracting event details into a DataFrame.
    
    Parameters:
    file_path (str): The path to the ICS file.
    
    Returns:
    pd.DataFrame: A DataFrame containing the event details, including 'summary', 'dtstart', 'dtend', 'location', and 'description'.
    """
    with open(file_path, 'r') as file:
        gcal = Calendar.from_ical(file.read())
        
    utc=pytz.UTC
    now = datetime.now().replace(tzinfo=utc)
    events = []
    
    for component in gcal.walk():
        if component.name == "VEVENT":
           # try:
                # Extract start time and ensure it's a datetime object with timezone info
                dtstart=component.get('dtstart').dt
                if isinstance(dtstart, datetime):
                    if not dtstart.tzinfo:
                        dtstart = dtstart.replace(tzinfo=pytz.UTC)
                else:  # if dtstart is a date object
                    dtstart = datetime.combine(dtstart, time.min, tzinfo=pytz.UTC)
                # Skip events that are in the past
                if dtstart < now:
                    continue
                
                # Extract end time and ensure it's a datetime object with timezone info
                dtend = component.get('dtend').dt
                if isinstance(dtend, datetime):
                    if not dtend.tzinfo:
                            dtend = dtend.replace(tzinfo=pytz.UTC)
                else:  # if dtend is a date object
                    dtend = datetime.combine(dtend, time.max, tzinfo=pytz.UTC)
                

                event = {
                    'summary': component.get('summary'),
                    'dtstart': dtstart,
                    'dtend': dtend,
                    'location': component.get('location'),
                    'description': component.get('description')
                }
                events.append(event)
            #except:
            #    print("Data ingnored: some error occured in:", component)

        # Filter events to only include those after now
    return pd.DataFrame(events)


def create_merged_cell_lookup(sheet) -> dict:
    """
    Creates a lookup dictionary for merged cells in a given sheet.
    
    This function iterates through all merged cell ranges in the given sheet,
    and creates a dictionary where the keys are the merged cell ranges (as strings)
    and the values are the values of the top-left cell in each merged cell range.

    Args:
        sheet (openpyxl.worksheet.worksheet.Worksheet): The worksheet object to process.

    Returns:
        dict: A dictionary with merged cell ranges as keys and the top-left cell values as values.
    """
    merged_lookup = {}
    for cell_group in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(str(cell_group))
        #if min_col == max_col:
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        merged_lookup[str(cell_group)] = top_left_cell_value
    return merged_lookup

 
def unmerge_cell_copy_top_value(workbook_path: str, output_save="", verbose: bool=False):
    """
    Unmerges cells in the given workbook and copies the top-left cell value to all cells in each previously merged range.
    
    This function opens the workbook at the given path, processes each worksheet by unmerging all merged cells,
    and copies the value of the top-left cell in each merged range to all cells in that range. The modified workbook
    is saved as "ready4Import.xlsx" in the current working directory.

    Args:
        workbook_path (str): The path to the Excel workbook to process.
        output_save (str): if not empty, will save the file to requested ouput.
        verbose (bool): If True, print debug information during processing. Default is False.

    Returns:
        openpyxl.workbook.workbook.Workbook: The modified workbook object.
    """
    wbook = openpyxl.load_workbook(workbook_path, data_only=True)
    
    for sheet in wbook.worksheets:
        lookup = create_merged_cell_lookup(sheet)
        if verbose: print(lookup)
        cell_group_list = lookup.keys()
        for cell_group in cell_group_list:
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(str(cell_group))
            sheet.unmerge_cells(str(cell_group))
            if verbose: print(min_col, min_row, max_col, max_row)
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                if verbose :print(lookup[cell_group])
                for cell in row:
                    cell.value = lookup[cell_group]
                    if verbose: print(cell.coordinate)
    if output_save : 
        wbook.save(output_save)
    return wbook

def search_string_in_workbook(wbook: "openpyxl workbook", search_string: str):
    """
    Searches for a specified string in all sheets of a workbook and extracts their coordinates along with sheet names and cell values.

    This function iterates through each worksheet in the provided openpyxl workbook, searches for the specified string in all cells,
    and stores the results, including the sheet name, cell coordinate, and cell value.

    Args:
        wbook (openpyxl.workbook.workbook.Workbook): The openpyxl workbook object to process.
        search_string (str): The string to search for in the workbook.

    Returns:
        list: A list of lists, where each inner list contains the sheet name, cell coordinate, and cell value of a match.
    """
    search_results=[]
    for sheet in wbook.worksheets:
        search_results+=search_string_in_worksheet(sheet,search_string)
    
    return search_results


def search_string_in_worksheet(sheet, search_string):
    search_results=[]
    for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if search_string in cell.value:
                        search_results.append([sheet.title,cell.coordinate, cell.value])
    return search_results

def extract_schedule_by_group_EI1(file_path : str, 
                                  course_name: str,
                                  group_name:str ,
                                  course_type: str, 
                                  display_group_schedule: bool=False):
    """
    Extracts the schedule of a specified course for a specified group from a given EI1 at ECN course schedule Excel file.
    
    Parameters:
    file_path (str): The path to the EI1 course schedule Excel file.
    course_name (str): The short name of the course to extract (e.g., 'FLUID').
    group_name (str): The name of the group to extract the schedule for.
    course_type (str): The type of course (e.g., 'TP', 'TD', 'CM').
    display_group_schedule (bool): If True, prints the schedule for the group. Defaults to False.
    
    Returns:
    list: A list of dictionaries containing the event details, with keys 'summary', 'dtstart', and 'dtend'.
    """
    class_slots={"M1": ["08:00","10:00"],"M2": ["10:15","12:15"],"S1": ["13:45","15:45"],"S2": ["16:00","18:00"]}
    tz = pytz.timezone('Europe/Paris')

    events=[]
    wbook=unmerge_cell_copy_top_value(file_path)
    
    #finding group row using column B
    for row in wbook.worksheets[0].iter_rows(min_col=2, min_row=4, max_col=2, max_row=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if group_name in cell.value:
                    groups_row=str(cell.row)
                    
    occurences=search_string_in_workbook(wbook , search_string=course_name)
    for occ in occurences:
        if occ[1][1:] != groups_row or occ[2].split()[1][:2] not in course_type: continue # Beware very hacky test
        # extracting date
        date_pos=occ[1][0]+"3"
        date=wbook[occ[0]][date_pos].value.split()[1] # dropping day name
        slot_pos=occ[1][0]+"4"
        slot=wbook[occ[0]][slot_pos].value.strip()
        
        dtstart= tz.localize(datetime.strptime(date+" "+class_slots[slot][0],'%d/%m/%y %H:%M'))#converting to date object
        dtend= tz.localize(datetime.strptime(date+" "+class_slots[slot][1],'%d/%m/%y %H:%M'))#converting to date object

        if display_group_schedule: 
            print(f"{wbook[occ[0]][date_pos].value.split()[0]:<8} {date} {slot} {occ[2]}\tGrp {group_name}")
            
        events.append({ 'summary': occ[2]+f" Grp {group_name}",
                        'dtstart': dtstart,
                        'dtend': dtend})
    return events

def extract_schedule_1sheet_format(file_path, 
                                   course_name, 
                                   date_column, 
                                   line_slot="2", 
                                   group_name=None, 
                                   course_type=None, 
                                   display_group_schedule=False):
    """
    Extracts the schedule for a specific course from a single-sheet formatted course schedule Excel file.
    
    Args:
        file_path (str): The path to the Excel file containing the course schedule.
        course_name (str): The name of the course to search for in the schedule.
        date_column (str): The column containing the dates in the schedule.
        line_slot (str, optional): The row number in the sheet to look for class slots (default is "2").
        group_name (str, optional): The group name (not used in current implementation).
        course_type (str, optional): The type of the course (not used in current implementation).
        display_group_schedule (bool, optional): If True, displays the schedule for the group (default is False).

    Returns:
        list: A list of dictionaries, each representing an event with keys 'summary', 'dtstart', and 'dtend'.
    
    Notes:
        This function is specific to the design of one-sheet course schedule xlsx files at ECN. 
        It assumes that the schedule is structured in a specific way with class slots defined as:
            - M1: 08:00-10:00
            - M2: 10:15-12:15
            - S1: 13:45-15:45
            - S2: 16:00-18:00
        The function handles merged cells and extracts the necessary information based on the course name.

        For successive courses, you can add a blank space at the end of the course name in order to disentangle these.
    """
    class_slots={"M1": [[ 8, 0],[10, 0]],
                 "M2": [[10,15],[12,15]],
                 "S1": [[13,45],[15,45]],
                 "S2": [[16, 0],[18, 0]]}
    tz = pytz.timezone('Europe/Paris')
    jour_to_dayshift={"Lundi":0, "Mardi":1, "Mercredi":2, "Jeudi":3, 'Vendredi':4}

    events=[]
    wbook=unmerge_cell_copy_top_value(file_path)
                    
    occurences=search_string_in_workbook(wbook, search_string=course_name)
    for occ in occurences:
        column_index="".join(filter(lambda x: x.isalpha(), occ[1])) #extracting column index (hack, there's probably a builtin way)
        #if occ[2].split()[1][:2] not in course_type: continue # Beware very hacky test
        # extracting date
        date_pos=date_column+"".join(filter(lambda x: x.isdigit(), occ[1]))
        date=wbook[occ[0]][date_pos].value # getting monday date
        # updating date using slot_pos
        day_of_week=wbook[occ[0]][column_index+str(int(line_slot)-1)].value
        date=date+timedelta(days=jour_to_dayshift[day_of_week])
        
        #print(date_pos,slot_pos)
        slot_pos=column_index+line_slot
        slot=wbook[occ[0]][slot_pos].value.strip()
        dtstart= tz.localize(date.replace(hour=class_slots[slot][0][0],minute=class_slots[slot][0][1]))#converting to date object
        dtend=   tz.localize(date.replace(hour=class_slots[slot][1][0],minute=class_slots[slot][1][1]))#converting to date object

        if display_group_schedule: 
            print(date.strftime("%a %d/%m/%y"),f"{slot} {clean_text(occ[2])}")
            
        events.append({ 'summary': clean_text(occ[2]),
                        'dtstart': dtstart,
                        'dtend': dtend})
    return events


def find_conflicting_events(combined_df):
    """
    Identify conflicting events in a DataFrame and return a DataFrame containing these conflicts.
    
    Parameters:
    combined_df (pd.DataFrame): A DataFrame containing events with 'summary', 'dtstart', and 'dtend' columns.
    
    Returns:
    pd.DataFrame: A DataFrame containing conflicting events with details about both conflicting events.
    """
    
    # Ensure the DataFrame is sorted by the start date
    combined_df = combined_df.sort_values(by='dtstart').reset_index(drop=True)
    
    # Create a list to store conflicting events
    conflicts = []

    # Iterate through the sorted DataFrame to identify conflicts
    for i in range(1, len(combined_df)):
        current_event = combined_df.iloc[i]
        previous_event = combined_df.iloc[i - 1]

        if current_event['dtstart'] < previous_event['dtend']:
            conflicts.append({
                'event1_summary': previous_event['summary'],
                'event1_dtstart': previous_event['dtstart'],
                'event1_dtend': previous_event['dtend'],
                'event2_summary': current_event['summary'],
                'event2_dtstart': current_event['dtstart'],
                'event2_dtend': current_event['dtend']
            })

    # Convert the list of conflicts to a DataFrame
    conflicts_df = pd.DataFrame(conflicts)
    return conflicts_df

def clean_text(text):
    # Replace newline, tab, and carriage return with a space
    text = text.replace('\n', ' ').replace('\t', ' ').replace('\r', ' ')
    # Replace multiple spaces with a single space
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def df_to_ics(df, output_file):
    """
    Convert a DataFrame to an ICS (iCalendar) file.

    Parameters:
    df (pd.DataFrame): DataFrame containing event data with the following columns:
        - 'summary' (str): A brief description of the event.
        - 'dtstart' (datetime): The start datetime of the event.
        - 'dtend' (datetime): The end datetime of the event.
        - 'location' (str, optional): The location of the event.
        - 'description' (str, optional): A description of the event.
    output_file (str): The path to the output ICS file.
    """
    
    # Create a new calendar
    cal = Calendar()
    
    # Set timezone to Europe/Paris (CEST)
    tz = pytz.timezone('Europe/Paris')
    
    # Iterate through the DataFrame and add events to the calendar
    for index, row in df.iterrows():
        event = Event()
        event.add('summary', row['summary'])
        event.add('dtstart', row['dtstart'])
        event.add('dtend', row['dtend'])
        event.add('location', row.get('location', ''))
        event.add('description', row.get('description', ''))
        cal.add_component(event)
    
    # Write the calendar to the output file
    with open(output_file, 'wb') as f:
        f.write(cal.to_ical())